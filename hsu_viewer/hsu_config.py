import json
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


"""
TODO:

- create dataset class and move relevant methods to it
- make data name consistent ie product_group, datatype, etc
- add docstrings
- add try statement for opening config file

"""

DATA_COLUMNS = [
    "mineral_per",
    "chemistry_",
    "position_",
]

INDEX_COLUMNS = {
    "filename": str,
    "box_number": int,
    "row_number": int,
    "meter_from": float,
    "meter_to": float,
}

SKIP_COLUMNS = [
    "filename",
    "path",
    "csv_path",
    "info_line_number",
    "Hole_ID",
    "box_number",
    "row_number",
    "meter_from",
    "meter_to",
    "meter_start",
    "meter_end",
    "position_raw",
    "None",
]


class HSUConfig:
    def __init__(self, config_path: Path | str = None) -> None:
        config_path = (
            config_path if isinstance(config_path, Path) else Path(config_path)
        )
        self.hsu_config_path: Path = config_path
        self.hsu_config: dict = self._get_hsu_config()

    def __getitem__(self, key: str) -> dict:
        return self.hsu_config.get(key)

    def _get_hsu_config(self) -> dict:
        cwd = Path(__file__).parent

        try:
            with open(cwd.joinpath(self.hsu_config_path), "r") as f:
                return json.load(f)
        except FileNotFoundError:
            return {}

    def add_dataset(self, dataset_path: str) -> str:
        dataset_path = Path(dataset_path)
        dataset_name = dataset_path.name
        dataset_config_path = dataset_path.joinpath(f"{dataset_name}.cfg")

        spec_images = self._get_spec_image_data(dataset_path.joinpath("Core"))
        core_images = self._get_core_image_data(dataset_path.joinpath("Photo"))

        csv_files = list(dataset_path.glob("*_DATA.csv"))
        if len(csv_files) > 0:
            spec_data, csv_data = self._parse_csv_data(csv_files[0])

        config_data = {
            "path": dataset_path.as_posix(),
            "csv_data": {"path": csv_files[0].as_posix(), **csv_data},
            "data": {
                "Spectral Images": spec_images,
                "Spectral Data": spec_data,
                "Corebox Images": core_images,
            },
        }

        if self.hsu_config.get(dataset_name) and self.hsu_config[
            dataset_name
        ].get("geochem_path"):
            geochem_path = Path(
                self.hsu_config[dataset_name].get("geochem_path")
            )
            geochem_data = self._get_geochem_data(geochem_path.as_posix())
            config_data["data"]["Additional Data"] = {
                "Geochemistry": {
                    **geochem_data,
                }
            }
            self.hsu_config[dataset_name] = {
                "path": dataset_config_path.as_posix(),
                "geochem_path": geochem_path.as_posix(),
            }
        else:
            self.hsu_config[dataset_name] = {
                "path": dataset_config_path.as_posix()
            }

        with open(dataset_config_path, "w") as f:
            json.dump(config_data, f)

        self._save_hsu_config()

        return dataset_name

    def add_geochem(
        self,
        geochem_path: str,
    ) -> str:
        geochem_path = Path(geochem_path)
        dataset_name = geochem_path.name.replace(".xlsx", "")
        if "_Geochemistry" in dataset_name:
            dataset_name = dataset_name.replace("_Geochemistry", "")

        geochem_data = self._get_geochem_data(geochem_path.as_posix())

        if self.hsu_config.get(dataset_name):
            dataset_config_path = self.hsu_config[dataset_name]["path"]
            # TODO: add catch if cannot open config
            with open(dataset_config_path, "r") as f:
                dataset_config = json.load(f)
            dataset_config["data"]["Additional Data"] = {
                "Geochemistry": {
                    **geochem_data,
                }
            }
            with open(dataset_config_path, "w") as f:
                json.dump(dataset_config, f)
        else:
            dataset_config_path = geochem_path.parent.joinpath(
                f"{dataset_name}.cfg"
            )
            with open(dataset_config_path, "w") as f:
                json.dump(
                    {
                        "path": geochem_path.as_posix(),
                        "geochem_only": True,
                        "data": {
                            "Additional Data": {
                                "Geochemistry": {
                                    **geochem_data,
                                }
                            },
                        },
                    },
                    f,
                )

            self.hsu_config[dataset_name] = {
                "path": dataset_config_path.as_posix(),
                "geochem_path": geochem_path.as_posix(),
                "geochem_only": True,
            }

        self._save_hsu_config()

        return dataset_name

    def _get_geochem_data(self, geochem_path: str) -> dict:
        wb = load_workbook(filename=geochem_path)
        ws = wb["Geochemistry"]
        col_headers = []
        meter_start = None
        meter_end = None
        geochem_data = {}
        for col in ws.iter_cols(min_row=2, min_col=2, values_only=True):
            header = col[0]
            col_headers.append(header)
            col_data = [
                c for c in col if isinstance(c, int) or isinstance(c, float)
            ]
            if header == "depth_start":
                meter_start = col_data[0]
            elif header == "depth_end":
                meter_end = col_data[-1]
            elif header is not None:
                header_pts = header.split("_")
                mineral_name = " ".join(header_pts[:-1])
                mineral_unit = header_pts[-1]
                geochem_data[mineral_name] = {
                    "name": mineral_name,
                    "unit": mineral_unit,
                    "path": geochem_path,
                    "meter_start": meter_start,
                    "meter_end": meter_end,
                    "min_value": 0,
                    "max_value": max(col_data),
                }

        return geochem_data

    def _save_hsu_config(self) -> None:
        keys = list(self.hsu_config.keys())
        keys.sort()
        self.hsu_config = {key: self.hsu_config[key] for key in keys}

        with open(self.hsu_config_path, "w") as f:
            json.dump(self.hsu_config, f)

    def dataset_path(self, dataset: str) -> dict:
        return self.hsu_config[dataset].get("path")

    def datasets(self) -> list:
        if self.hsu_config:
            return list(self.hsu_config.keys())
        else:
            return []

    def _get_spec_image_data(self, dataset_path: Path) -> list:
        spec_im_dict = {}
        if dataset_path.is_dir():
            for path in dataset_path.iterdir():
                if path.is_dir() and "mask_" not in path.name:
                    dir_info = path.name.split("_")
                    image_type = dir_info[0].title()
                    name = " ".join(dir_info[1:]).title()

                    if name:
                        meta_data = {
                            "name": name,
                            "path": path.as_posix(),
                        }

                        if spec_im_dict.get(image_type):
                            spec_im_dict[image_type][name] = meta_data
                        else:
                            spec_im_dict[image_type] = {name: meta_data}

        return spec_im_dict

    def _get_core_image_data(self, dataset_path: Path) -> list:
        core_im_dict = {}
        if dataset_path.is_dir():
            for path in dataset_path.iterdir():
                if path.is_dir():
                    meta_data = {"name": path.name, "path": path.as_posix()}
                    core_im_dict[path.name] = {path.name: meta_data}
        return core_im_dict

    def _parse_csv_data(self, csv_path: Path) -> list:
        csv_data_dict = {}
        spectral_data = {}
        csv_data = np.genfromtxt(
            csv_path, delimiter=",", max_rows=5, dtype="str", comments=None
        )
        columns = csv_data[1]
        indexers = {
            col: idx
            for idx, col in enumerate(columns)
            if col in list(INDEX_COLUMNS.keys())
        }

        meter_from_cols = (
            indexers["meter_from"]
            if isinstance(indexers["meter_from"], list)
            else [indexers["meter_from"]]
        )

        meter_to_cols = (
            indexers["meter_to"]
            if isinstance(indexers["meter_to"], list)
            else [indexers["meter_to"]]
        )

        meter_warning = None
        meter_data = np.genfromtxt(
            csv_path,
            delimiter=",",
            dtype="float",
            comments=None,
            skip_header=5,
            usecols=[*meter_from_cols, *meter_to_cols],
        )
        if meter_data.max() < 9999:
            meter_start = meter_data.min()
            meter_end = meter_data.max()
        else:
            meter_start = 0
            meter_end = meter_data.shape[0]
            meter_warning = True

        csv_data_dict = {
            **csv_data_dict,
            **indexers,
            "meter_start": meter_start,
            "meter_end": meter_end,
            "n_rows": meter_data.shape[0],
        }

        if meter_warning:
            csv_data_dict["meter_missing"] = True

        for idx, col in enumerate(columns):
            if (
                any(d in col.lower() for d in DATA_COLUMNS)
                and col.lower() not in SKIP_COLUMNS
            ):
                meter_idx = np.searchsorted(meter_from_cols, [idx], "left") - 1
                data_type = [dt for dt in DATA_COLUMNS if dt in col.lower()][0]
                name = col.replace(data_type, "").split("_")[1:]
                name = " ".join(name)

                meta_data = {
                    "name": name,
                    "unit": csv_data[2, idx],
                    "min_value": csv_data[3, idx],
                    "max_value": csv_data[4, idx],
                    "meter_from": meter_from_cols[meter_idx[0]],
                    "meter_to": meter_to_cols[meter_idx[0]],
                    "column": idx,
                }

                if data_type == "mineral_per":
                    data_type = "Mineral Percent"
                elif data_type == "chemistry_":
                    data_type = "Chemistry"
                elif data_type == "position_":
                    data_type = "Position"

                if spectral_data.get(data_type):
                    spectral_data[data_type][name] = meta_data
                else:
                    spectral_data[data_type] = {name: meta_data}

        return spectral_data, csv_data_dict
