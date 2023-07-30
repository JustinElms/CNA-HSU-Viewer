from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure
from matplotlib.ticker import NullFormatter
from PySide6.QtCore import Slot

from components.data_panel import DataPanel
from data.dataset import Dataset
from hsu_viewer.worker import Worker

"""
TODO
-resize plot instead of replotting it?
-add offset at top of panel
"""


class SpectralPlotPanel(DataPanel):
    def __init__(
        self,
        parent=None,
        threadpool=None,
        resolution: int = 0,
        dataset: Dataset = None,
        plot_colors: dict = None,
        **kwargs,
    ) -> None:
        super().__init__(
            parent=parent,
            resolution=resolution,
            dataset=dataset,
            threadpool=threadpool,
            **kwargs,
        )

        self.width = 180
        self.setFixedWidth(self.width)
        self.image_resolution = resolution
        self.depth = dataset.meter_end()

        self.setToolTip(f"{self.dataset_name} {self.data_name}")

        self.plot_colors = plot_colors

        self.loading.emit(True)
        self.get_plot()

    def get_plot(self) -> None:
        if self.threadpool:
            if self.data_subtype == "Geochemistry":
                worker = Worker(self._load_geochem_data)
            else:
                worker = Worker(self._load_spectral_data)
            worker.signals.result.connect(self._plot_spectral_data)
            worker.signals.finished.connect(self._on_finish)
            self.threadpool.start(worker)
        else:
            if self.data_subtype == "Geochemistry":
                result = self._load_geochem_data()
            else:
                result = self._load_spectral_data()
            self._plot_spectral_data(result)
            self.loading.emit(False)

    def _load_spectral_data(self) -> None:
        csv_path = Path(self.csv_data.get("path"))

        data = np.genfromtxt(
            csv_path,
            delimiter=",",
            dtype="float",
            comments=None,
            skip_header=5,
            usecols=[
                self.dataset_info.get("meter_from"),
                self.dataset_info.get("meter_to"),
                self.dataset_info.get("column"),
            ],
        )

        if data[-1, 1] >= 9999:
            data[:, 0] = np.arange(0, data.shape[0], 1)
            data[:, 1] = np.arange(1, data.shape[0] + 1, 1)

        if data[0, 0] != 0:
            data = np.insert(
                data.astype(float), 0, [0, data[0, 0], np.nan], axis=0
            )

        bar_widths = data[:, 1] - data[:, 0]
        bar_centers = (data[:, 0] + data[:, 1]) / 2
        meter_start = data[0, 0]
        meter_end = data[-1, 1]
        spectral_data = data[:, 2]

        return bar_widths, bar_centers, meter_start, meter_end, spectral_data

    def _load_geochem_data(self) -> None:
        geochem_path = self.dataset.geochem_path(self.data_name)

        wb = load_workbook(filename=geochem_path)
        ws = wb["Geochemistry"]
        meter_start = None
        meter_end = None
        mineral_unit = self.dataset.data(
            self.data_type, self.data_subtype, self.data_name
        ).get("unit")

        for col in ws.iter_cols(min_row=2, min_col=2, values_only=True):
            header = col[0]
            col_data = [
                c for c in col if isinstance(c, int) or isinstance(c, float)
            ]
            if header == "depth_start":
                meter_start = np.array(col_data)
            elif header == "depth_end":
                meter_end = np.array(col_data)
            elif header == f"{self.data_name}_{mineral_unit}":
                data = np.array(col_data)

        if meter_end[-1] >= 9999:
            meter_start = np.arange(0, data.shape[0], 1)
            meter_end = np.arange(1, data.shape[0] + 1, 1)

        if meter_start[0] != 0:
            meter_start = np.insert(meter_start, 0, 0)
            meter_end = np.insert(meter_end, 0, meter_start[0])
            data = np.insert(data.astype(float), 0, np.nan)

        bar_widths = meter_end - meter_start
        bar_centers = (meter_end + meter_start) / 2
        meter_start = meter_start[0]
        meter_end = meter_end[-1]

        return bar_widths, bar_centers, meter_start, meter_end, data

    def _plot_spectral_data(self, result: tuple) -> None:
        # create plot figure and canvas
        bar_widths, bar_centers, meter_start, meter_end, spectral_data = result

        if self.layout.count() > 0:
            self.layout.itemAt(0).widget().deleteLater()

        plot_color = self.plot_colors.get(self.data_name)
        height = (meter_end - meter_start) * self.resolution
        plot_fig = Figure(
            figsize=(self.width / 100, height / 100),
            dpi=100,
            facecolor="#000000",
        )
        plotCanvas = FigureCanvasQTAgg(plot_fig)
        plotCanvas.draw()
        plotCanvas.setMouseTracking(True)

        # # make plot tooltip to display cursor coordinates and plot info
        # plotCanvas.mouseMoveEvent = lambda event: self.makePlotToolTip(
        #     event,
        #     newDataWindow,
        #     0,
        #     axis_max,
        #     plotDepth[0],
        #     plotDepth[-1],
        #     unit,
        #     "spec",
        #     " ",
        # )

        try:
            axis_min = float(self.dataset_info.get("min_value"))
            axis_max = float(self.dataset_info.get("max_value"))
        except ValueError:
            if self.data_subtype == "Position":
                [min, max] = self.data_name.split(" ")
                axis_min = float(min)
                axis_max = float(max)
            else:
                axis_min = spectral_data.min()
                axis_max = spectral_data.max()

        plot_fig.clear()
        plot = plot_fig.add_axes([0, 0, 1, 1])
        plot.barh(
            bar_centers,
            spectral_data,
            bar_widths,
            color=plot_color,
        )
        plot.set_xticks([axis_min, (axis_max + axis_min) / 2, axis_max])
        plot.set_ylim(meter_end, meter_start)
        plot.set_xlim(axis_min, axis_max)
        plot.set_frame_on(False)
        plot.grid(color="#323232")
        plot.xaxis.set_major_formatter(NullFormatter())
        plot.yaxis.set_major_formatter(NullFormatter())
        plotCanvas.draw()

        self.insert_plot(plotCanvas)

    @Slot(int)
    def zoom_changed(self, resolution: int) -> None:
        self.loading.emit(True)
        self.resolution = resolution
        self.get_plot()

    def insert_plot(self, plot: FigureCanvasQTAgg) -> None:
        if self.layout.count() == 0:
            self.layout.addWidget(plot)
            self.layout.addStretch()
        else:
            self.layout.insertWidget(self.layout.count() - 1, plot)
